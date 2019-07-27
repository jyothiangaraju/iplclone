import click
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from copy import copy
import sys; print('Python %s on %s' % (sys.version, sys.platform))
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE','appsday.settings')
django.setup()
from onlineapp.models import *


@click.group()
def cli():
    return


@cli.command()
def importseason():
    "concatenates passed in strings with delimiter"
    #cr = db.cursor()
    wb1 = load_workbook(filename="matches.xlsx")
    ws=wb1['worksheet']
    flag=0
    for row in  ws:
        if (flag == 0):
            flag = 1
            continue
        row_data=[]
        for col in row:
            if (type("str").__name__=='str'):
                 row_data.append(col.value.lower())
            else:
                row_data.append(col.value)
        c1=Season(season=row_data[1],city=row_data[2],team1=row_data[4],team2=row_data[5],toss_winner=row_data[6],toss_decision=row_data[7],result=row_data[8],dl_applied=row_data[9],winner=row_data[10],winner_by_runs=row_data[11],winner_by_wickets=row_data[12],player_of_the_match=row_data[13],venue=row_data[14],umpire1=row_data[15],umpire2=row_data[16],umpire3=row_data[17])
        c1.save()
        break

# @cli.command()
# def  importstudents():
#     wb1 = load_workbook(filename="students.xlsx")
#     ws = wb1['Current']
#     flag = 0
#     for row in ws:
#         if(flag==0):
#             flag=1
#             continue
#         row_data = []
#         for col in row:
#             row_data.append(col.value.lower())
#         print(row_data)
#         c1 = Student(name=row_data[0], email=row_data[2], db_folder=row_data[3], college_id=row_data[1])
#         c1.save()
#
# @cli.command()
# def importmarks():
#     wb1 = load_workbook(filename="results.xlsx")
#     ws = wb1['result']
#     flag = 0
#     for row in ws:
#         if (flag == 0):
#             flag = 1
#             continue
#         row_data = []
#         num=0
#         for col in row:
#             if (num == 1):
#                 i = col.value[7:].find('_')
#                 row_data.append(col.value[8 + i:-5])
#             else:
#                 row_data.append(col.value[1:])
#             num = num + 1
#         c1 = MockTest1(student_id=row_data[1],problem1=row_data[2], problem2=row_data[3], problem3=row_data[4],problem4=row_data[5],total=row_data[6])
#         c1.save()

if __name__=="__main__":
      cli()