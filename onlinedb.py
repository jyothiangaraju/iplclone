import click
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from copy import copy
import sys; print('Python %s on %s' % (sys.version, sys.platform))
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE','appsday.settings')
django.setup()
from onlineapp.models import *


@click.group()
def cli():
    return


@cli.command()
def importseason():
    "concatenates passed in strings with delimiter"
    #cr = db.cursor()
    wb1 = load_workbook(filename="matches.xlsx")
    ws=wb1['Worksheet']
    flag=0
    for row in  ws:
        if (flag == 0):
            flag = 1
            continue
        row_data=[]
        for col in row:
            if (type(col.value).__name__=='str'):
                 row_data.append(col.value.lower())
            else:
                row_data.append(col.value)
        if(row_data[1]==2008):
            c1=Season(season=row_data[1],city=row_data[2],team1=row_data[4],team2=row_data[5],toss_winner=row_data[6],toss_decision=row_data[7],result=row_data[8],dl_applied=row_data[9],winner=row_data[10],winner_by_runs=row_data[11],winner_by_wickets=row_data[12],player_of_the_match=row_data[13],venue=row_data[14],umpire1=row_data[15],umpire2=row_data[16],umpire3=row_data[17])
            c1.save()

@cli.command()
def importmatch():
    "concatenates passed in strings with delimiter"
    #cr = db.cursor()
    wb1 = load_workbook(filename="deliveries.xlsx")
    ws=wb1['Worksheet']
    flag=0
    for row in  ws:
        if (flag == 0):
            flag = 1
            continue
        row_data=[]
        for col in row:
            if (type(col.value).__name__=='str'):
                 row_data.append(col.value.lower())
            else:
                row_data.append(col.value)
        print(1337+row_data[0])
        c1=Match(match_id=1337+row_data[0],inning=row_data[1],batting_team=row_data[2],bowling_team=row_data[3],over=row_data[4],ball=row_data[5],batsman=row_data[6],non_striker=row_data[7],bowler=row_data[8],is_super_over=row_data[9],wide_runs=row_data[10],bye_runs=row_data[11],legbye_runs=row_data[12],noball_runs=row_data[13],penalty_runs=row_data[14],batsman_runs=row_data[15],extra_runs=row_data[16],total_runs=row_data[17],player_dismissed=row_data[18],dismissal_kind=row_data[19],fielder=row_data[20])
        c1.save()



if __name__=="__main__":
      cli()